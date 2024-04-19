from models import *
import pandas as pd
import datetime
import streamlit as st
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import io


st.header("Procolo de SCs")
st.markdown("""
        <p>⚠ Esta tela permite gerar uma planilha-protocolo contendo todas as solicitações já registradas para visualização e análise. 
            Você pode selecionar a data de início e término para filtrar as solicitações de acordo com o dia de atendimento.<p>
""", unsafe_allow_html=True)


# Carregar a planilha existente
planilha_template = 'template_excel.xlsx'
wb = load_workbook(planilha_template)
ws = wb['MAR']

# Limpar os dados existentes
max_row = ws.max_row
start_row = 10
num_cols = 10
for i in range(start_row, max_row + 1):
    for j in range(2, num_cols + 2):
        ws.cell(row=i, column=j).value = ''

# Carregar os dados em um DataFrame
lojas = [lj for lj in col_filiais.find()]
start_date = st.date_input("Selecione a data de início:",  pd.to_datetime('today'), format='DD/MM/YYYY')
end_date = st.date_input("Selecione a data de término:", pd.to_datetime('today'), format='DD/MM/YYYY')

# Formatar as datas para o formato esperado na consulta MongoDB
start_date_format = start_date.strftime("%d/%m/%Y") + " 00:00:00"
end_date_format = end_date.strftime("%d/%m/%Y") + " 23:59:59"

# Filtrar os dados por data de abertura
query = {'data_atendimento': {'$gte': start_date_format, '$lte': end_date_format}}
data = [dados for dados in col_solicitacao.find(query)]



df = pd.DataFrame(data)

# Mapear as colunas do DataFrame para as colunas da planilha
column_mapping = {
    'Nº Solicitação': 'nr_solicitacao',
    'Cod. Filial Senior': 'cod_loja',
    'Loja': 'loja',
    'REGIONAL': 'MA',
    'Data Emissão': 'data_abertura',
    'FORNECEDOR': 'forncedor',
    'OC': 'oc',
    'Tipologia': 'tipologia',
    'Status': 'N'
}

# Preencher a planilha com os dados
start_row = 10
for index, row in df.iterrows():
    for col_planilha, col_df in column_mapping.items():
        if col_df in row:
            valor_celula = row[col_df]
            if valor_celula != '':
                col_index = list(column_mapping.keys()).index(col_planilha) + 2
                ws.cell(row=start_row + index, column=col_index, value=valor_celula)

# Preencher a coluna "REGIONAL" com base na UF da loja
lojas_map = {loja['nome_loja']: loja['uf'] for loja in lojas}
for index, row in df.iterrows():
    loja_nome = row['loja']
    if loja_nome in lojas_map:
        regional_value = lojas_map[loja_nome]
        regional_col = list(column_mapping.keys()).index('REGIONAL') + 2
        ws.cell(row=start_row + index, column=regional_col, value=regional_value)

# Aplicar estilo de borda apenas nas células preenchidas com dados
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(data) - 1,
                        min_col=2, max_col=num_cols + 1):
    for cell in row:
        if cell.value:
            cell.border = thin_border

# Remover linhas de grade extras
for row in ws.iter_rows(min_row=start_row + len(data), max_row=ws.max_row,
                        min_col=2, max_col=num_cols + 1):
    for cell in row:
        cell.border = None

# Definir a área de impressão automaticamente
ws.print_area = ws.dimensions

# Criar um buffer de bytes para armazenar o arquivo Excel
excel_buffer = io.BytesIO()

# Salvar a planilha no buffer
wb.save(excel_buffer)

# Disponibilizar o arquivo Excel para download
excel_buffer.seek(0)
st.download_button(
    label="Baixar Excel",
    data=excel_buffer,
    file_name=f'solicitacoes_{start_date_format}.xlsx',
    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)

print("Protocolo de SC Gerado")
